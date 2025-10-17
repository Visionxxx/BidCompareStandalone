from __future__ import annotations

import base64
import io
import re
import xml.etree.ElementTree as ET
from datetime import datetime
from typing import Any

import pandas as pd
import numpy as np
from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

app = FastAPI(title="Bid Compare Tool", version="1.0.0")

# CORS configuration
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # In production, specify actual origins
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    cols = {c.lower(): c for c in df.columns}

    def pick(*names: str, fallback: str | None = None) -> str | None:
        for name in names:
            if name in cols:
                return cols[name]
        return fallback

    post_col = pick("postnr") or df.columns[0]
    desc_col = pick("beskrivelse", "description") or (df.columns[1] if len(df.columns) > 1 else post_col)
    unit_col = pick("enhet", "unit") or (df.columns[2] if len(df.columns) > 2 else desc_col)
    qty_col = pick("mengde", "qty") or (df.columns[3] if len(df.columns) > 3 else unit_col)
    price_col = pick("pris", "unit_price")
    sum_col = pick("sum", "sum_amount")
    code_col = pick("kode", "nskode", "ns_code", "code")

    use_cols = [post_col, desc_col, unit_col, qty_col]
    if price_col:
        use_cols.append(price_col)
    if sum_col:
        use_cols.append(sum_col)
    if code_col:
        use_cols.append(code_col)

    df2 = df[use_cols].copy()
    target_cols = ["postnr", "beskrivelse", "enhet", "qty"]
    if price_col:
        target_cols.append("unit_price")
    if sum_col:
        target_cols.append("sum_amount")
    if code_col:
        target_cols.append("ns_code")
    df2.columns = target_cols

    if "qty" in df2.columns:
        df2["qty"] = pd.to_numeric(df2["qty"], errors="coerce").fillna(0.0)

    if "unit_price" in df2.columns:
        df2["unit_price"] = (
            df2["unit_price"]
            .astype(str)
            .str.replace(" ", "", regex=False)
            .str.replace(",", ".", regex=False)
        )
        df2["unit_price"] = pd.to_numeric(df2["unit_price"], errors="coerce").fillna(0.0)

    if "sum_amount" in df2.columns:
        df2["sum_amount"] = (
            df2["sum_amount"]
            .astype(str)
            .str.replace(" ", "", regex=False)
            .str.replace(",", ".", regex=False)
        )
        df2["sum_amount"] = pd.to_numeric(df2["sum_amount"], errors="coerce").fillna(0.0)
    else:
        df2["sum_amount"] = df2.get("qty", 0.0) * df2.get("unit_price", 0.0)

    if "ns_code" not in df2.columns:
        df2["ns_code"] = ""
    if "ns_title" not in df2.columns:
        df2["ns_title"] = ""
    if "specification" not in df2.columns:
        df2["specification"] = df2.get("beskrivelse", "").astype(str)
    if "is_option" not in df2.columns:
        df2["is_option"] = False
    if "kapittel_navn" not in df2.columns:
        df2["kapittel_navn"] = ""

    def chapter_of(value: Any) -> str:
        text = str(value or "").strip()
        return text[:2] if len(text) >= 2 else "00"

    df2["kapittel"] = df2["postnr"].apply(chapter_of)
    df2["ns_title"] = df2.get("ns_title", "").astype(str)
    df2["specification"] = df2.get("specification", df2.get("beskrivelse", "")).astype(str)
    df2["kapittel_navn"] = df2.get("kapittel_navn", "").astype(str)
    return df2


def _read_tabular(name: str, data: bytes) -> pd.DataFrame:
    try:
        if name.lower().endswith(".xlsx") or name.lower().endswith(".xls"):
            return pd.read_excel(io.BytesIO(data))
        try:
            df = pd.read_csv(io.BytesIO(data), sep=";")
            if df.shape[1] == 1:
                df = pd.read_csv(io.BytesIO(data))
        except Exception:
            df = pd.read_csv(io.BytesIO(data), encoding="latin-1")
        return df
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Could not read {name}: {exc}") from exc


def _to_float(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return 0.0
    text = text.replace(" ", "").replace("\u00a0", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def _parse_ns3459_xml(data: bytes, name: str) -> pd.DataFrame:
    try:
        root = ET.fromstring(data)
    except ET.ParseError as exc:
        raise HTTPException(status_code=400, detail=f"Could not read {name}: {exc}") from exc

    if root.tag.startswith("{"):
        ns_uri = root.tag.split("}")[0][1:]
    else:
        ns_uri = ""

    def tag(label: str) -> str:
        return f"{{{ns_uri}}}{label}" if ns_uri else label

    chapter_names: dict[str, str] = {}
    postnrplan = root.find(f"{tag('Pristilbud')}/{tag('ProsjektNS')}/{tag('Postnrplan')}")
    if postnrplan is not None:
        for element in postnrplan.findall(f".//{tag('PostnrdelKode')}"):
            type_value = (element.findtext(tag("Type")) or "").strip()
            if type_value != "Type1":
                continue
            kode = (element.findtext(tag("Kode")) or "").strip()
            navn = (element.findtext(tag("Navn")) or "").strip()
            if kode and navn and kode not in chapter_names:
                chapter_names[kode] = navn

    posts = root.findall(f".//{tag('Post')}")
    if not posts:
        raise HTTPException(status_code=400, detail=f"No posts found in {name}")

    records: list[dict[str, Any]] = []
    for post in posts:
        postnr = (post.findtext(tag("Postnr")) or "").strip()
        tekst_main = post.findtext(f"{tag('Tekst')}/{tag('Uformatert')}") or ""
        tekst_main = (
            tekst_main.replace("\r\n", "\n").replace("\r", "\n").strip()
        )

        prisinfo = post.find(tag("Prisinfo"))
        enhet = ""
        qty = 0.0
        unit_price = 0.0
        sum_amount = 0.0
        if prisinfo is not None:
            enhet = (prisinfo.findtext(tag("Enhet")) or "").strip()
            qty = _to_float(prisinfo.findtext(tag("Mengde")))
            unit_price = _to_float(prisinfo.findtext(tag("Enhetspris")))
            sum_amount = _to_float(prisinfo.findtext(tag("Sum")))
        if sum_amount == 0.0:
            sum_amount = qty * unit_price

        kapittel = ""
        for pn in post.findall(f"{tag('Postnrdeler')}/{tag('Postnrdel')}"):
            if (pn.findtext(tag("Type")) or "").strip() == "Type1":
                kapittel = (pn.findtext(tag("Kode")) or "").strip()
                break
        if not kapittel and postnr:
            kapittel = postnr.split(".")[0]

        kapittel_navn = chapter_names.get(kapittel, "")
        def parse_option(value: Any) -> bool:
            text = str(value or "").strip().lower()
            if text in {"true", "1", "ja", "yes"}:
                return True
            return False

        is_option = bool(prisinfo is not None and parse_option(prisinfo.get("Opsjon")))

        kode = post.find(tag("Kode"))
        ns_code = ""
        ns_title = ""
        if kode is not None:
            ns_code = (kode.findtext(tag("ID")) or "").strip()
            ns_title = (kode.findtext(f"{tag('Kodetekst')}/{tag('Overskrift')}") or "").strip()
            kodetekst = kode.find(tag("Kodetekst"))
        else:
            kodetekst = None

        spec_parts: list[str] = []
        seen = set()

        def add_part(val: str):
            val = val.strip()
            if not val:
                return
            key = val.lower()
            if key in seen:
                return
            seen.add(key)
            spec_parts.append(val)

        add_part(ns_title)
        add_part(tekst_main)

        if kodetekst is not None:
            for u in kodetekst.findall(f".//{tag('Uformatert')}"):
                add_part(u.text or "")
            for txt in kodetekst.findall(f".//{tag('Tekst')}"):
                if txt.get("OriginalFormat", "") == "RTF":
                    continue
                inner = txt.findtext(tag("Uformatert"))
                if inner:
                    add_part(inner)
                add_part(txt.text or "")
        else:
            for u in post.findall(f".//{tag('Uformatert')}"):
                add_part(u.text or "")
            for txt in post.findall(f".//{tag('Tekst')}"):
                if txt.get("OriginalFormat", "") == "RTF":
                    continue
                add_part(txt.text or "")

        if not spec_parts and tekst_main:
            add_part(tekst_main)

        specification = "\n\n".join(spec_parts)
        short_description = ns_title or (spec_parts[0] if spec_parts else tekst_main)

        records.append(
            {
                "postnr": postnr,
                "beskrivelse": short_description,
                "enhet": enhet,
                "qty": qty,
                "unit_price": unit_price,
                "sum_amount": sum_amount,
                "kapittel": kapittel,
                "kapittel_navn": kapittel_navn,
                "ns_code": ns_code,
                "ns_title": ns_title,
                "specification": specification,
                "is_option": is_option,
            }
        )

    return pd.DataFrame.from_records(
        records,
        columns=[
            "postnr",
            "beskrivelse",
            "enhet",
            "qty",
            "unit_price",
            "sum_amount",
            "kapittel",
            "kapittel_navn",
            "ns_code",
            "ns_title",
            "specification",
            "is_option",
        ],
    )


def _extract_company_name(data: bytes) -> str:
    try:
        root = ET.fromstring(data)
    except ET.ParseError:
        return ""

    ns_uri = ""
    if root.tag.startswith("{"):
        ns_uri = root.tag.split("}")[0][1:]

    def tag(label: str) -> str:
        return f"{{{ns_uri}}}{label}" if ns_uri else label

    paths = [
        f"{tag('Pristilbud')}/{tag('Generelt')}/{tag('Avsender')}/{tag('Firma')}/{tag('Navn')}",
        f"{tag('Prisforesporsel')}/{tag('Generelt')}/{tag('Avsender')}/{tag('Firma')}/{tag('Navn')}",
        f"{tag('ProsjektNS')}/{tag('Generelt')}/{tag('Avsender')}/{tag('Firma')}/{tag('Navn')}",
    ]
    for path in paths:
        value = root.findtext(path)
        if value:
            return value.strip()
    return ""


def _to_records(df: pd.DataFrame) -> list[dict[str, Any]]:
    return df.replace({pd.NA: None}).to_dict(orient="records")


def _collect_chapter_titles(bids: dict[str, pd.DataFrame]) -> dict[str, str]:
    titles: dict[str, str] = {}

    def normalize(text: str) -> str:
        collapsed = re.sub(r"\s+", " ", text.strip())
        collapsed = collapsed.rstrip(" .;-")
        if not collapsed:
            return ""
        letters = [ch for ch in collapsed if ch.isalpha()]
        if letters:
            upper_ratio = sum(ch.isupper() for ch in letters) / len(letters)
        else:
            upper_ratio = 0.0
        if upper_ratio >= 0.6:
            lowered = collapsed.lower()
            parts = re.split(r"([-/–—])", lowered)
            rebuilt = "".join(part.title() if idx % 2 == 0 else part for idx, part in enumerate(parts))
            collapsed = rebuilt
        return collapsed

    for df in bids.values():
        for _, row in df.iterrows():
            code = str(row.get("kapittel") or "").strip()
            if not code or code in titles:
                continue

            chapter_name = normalize(str(row.get("kapittel_navn") or ""))
            if chapter_name:
                titles[code] = chapter_name
                continue

            candidates: list[str] = []
            ns_title = str(row.get("ns_title") or "").strip()
            if ns_title:
                candidates.append(ns_title)

            specification = str(row.get("specification") or "").strip()
            if specification:
                spec_line = specification.splitlines()[0].strip()
                if spec_line:
                    candidates.append(spec_line)

            description = str(row.get("beskrivelse") or "").strip()
            if description:
                candidates.append(description)

            for candidate in candidates:
                if not candidate or candidate.lower() == "sum":
                    continue

                normalized = normalize(candidate)
                if not normalized:
                    continue

                text = normalized if len(normalized) <= 120 else f"{normalized[:117]}..."
                titles[code] = text
                break

    return titles


def _aggregate_bid_rows(df: pd.DataFrame, unit_label: str, sum_label: str) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for postnr, group in df.groupby("postnr", dropna=False):
        qty = pd.to_numeric(group.get("qty", pd.Series(dtype=float)), errors="coerce")
        unit_price = pd.to_numeric(group.get("unit_price", pd.Series(dtype=float)), errors="coerce")
        sum_amount = pd.to_numeric(group.get("sum_amount", pd.Series(dtype=float)), errors="coerce")

        mask = qty.notna() & unit_price.notna()
        if mask.any():
            qty_sum = float(qty[mask].sum())
            if qty_sum > 0:
                weighted_price = float((unit_price[mask] * qty[mask]).sum() / qty_sum)
            else:
                weighted_price = float(unit_price[mask].mean()) if unit_price[mask].notna().any() else None
        elif unit_price.notna().any():
            weighted_price = float(unit_price.mean(skipna=True))
        else:
            weighted_price = None

        total_sum = float(sum_amount.sum(skipna=True)) if sum_amount.notna().any() else None

        if weighted_price is None and total_sum is None:
            continue

        rows.append(
            {
                "postnr": str(postnr),
                unit_label: weighted_price,
                sum_label: total_sum,
            }
        )
    if not rows:
        return pd.DataFrame(columns=["postnr", unit_label, sum_label])
    return pd.DataFrame(rows)


def _lighten_hex(color: str, factor: float) -> str:
    color = color.lstrip("#")
    if len(color) != 6:
        return "FFFFFF"
    r = int(color[0:2], 16)
    g = int(color[2:4], 16)
    b = int(color[4:6], 16)
    r = int(r + (255 - r) * factor)
    g = int(g + (255 - g) * factor)
    b = int(b + (255 - b) * factor)
    return f"{max(0, min(255, r)):02X}{max(0, min(255, g)):02X}{max(0, min(255, b)):02X}"


def _build_matrix_excel(
    matrix: pd.DataFrame,
    sum_columns: list[str],
    unit_columns: list[str],
    column_colors: dict[str, str] | None = None,
) -> str:
    if matrix.empty:
        return ""

    buffer = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sammenligning per post"

    headers = list(matrix.columns)
    worksheet.append(headers)

    default_header_fill = PatternFill(fill_type="solid", start_color="E2E8F0", end_color="E2E8F0")
    header_font = Font(bold=True, color="0F172A")
    header_alignment = Alignment(horizontal="center", vertical="center")
    for cell in worksheet[1]:
        cell.font = header_font
        cell.alignment = header_alignment

    currency_columns = set(sum_columns + unit_columns + ["lavest_sum", "std_avvik", "snitt"])
    percent_columns = {"std_pct"}
    z_score_columns = {col for col in matrix.columns if col.endswith("(z-score)")}

    for record in matrix.to_dict(orient="records"):
        row_values: list[Any] = []
        for column in headers:
            value = record.get(column)
            if isinstance(value, np.generic):
                value = value.item()
            if value is None or (isinstance(value, float) and pd.isna(value)):
                row_values.append(None)
                continue
            if column in percent_columns and isinstance(value, (int, float)):
                row_values.append(float(value) / 100.0)
            else:
                row_values.append(value)
        worksheet.append(row_values)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    border_side = Side(style="thin", color="CBD5F5")
    table_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    right_alignment = Alignment(horizontal="right")
    percent_alignment = Alignment(horizontal="right")

    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = table_border

    for index, column in enumerate(headers, start=1):
        header_cell = worksheet.cell(row=1, column=index)
        if column_colors and column in column_colors:
            base = column_colors[column]
            if column in unit_columns:
                fill_color = _lighten_hex(base, 0.7)
            elif column in sum_columns:
                fill_color = _lighten_hex(base, 0.5)
            else:
                fill_color = base.lstrip("#").upper()
            header_cell.fill = PatternFill(fill_type="solid", start_color=fill_color, end_color=fill_color)
        else:
            header_cell.fill = default_header_fill

        column_letter = get_column_letter(index)
        max_length = 0
        for cell in worksheet[column_letter]:
            if cell.value is None:
                continue
            display_value = cell.value
            if column in percent_columns and cell.row > 1 and isinstance(display_value, (int, float)):
                display_value = f"{display_value * 100:.2f} %"
            length = len(str(display_value))
            max_length = max(max_length, length)
        worksheet.column_dimensions[column_letter].width = max(14, max_length + 2)

        if column in currency_columns:
            for cell in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=index, max_col=index):
                target = cell[0]
                if target.value is None:
                    continue
                target.number_format = '#,##0.00 "kr"'
                target.alignment = right_alignment
        elif column in percent_columns:
            for cell in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=index, max_col=index):
                target = cell[0]
                if target.value is None:
                    continue
                target.number_format = "0.00%"
                target.alignment = percent_alignment
        elif column in z_score_columns:
            for cell in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=index, max_col=index):
                target = cell[0]
                if target.value is None:
                    continue
                target.number_format = "0.00"
                target.alignment = right_alignment
        elif column != "postnr" and column != "vinner":
            for cell in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=index, max_col=index):
                target = cell[0]
                if target.value is None:
                    continue
                target.alignment = right_alignment

    bold_font = Font(bold=True)
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        if str(row[0].value).strip().upper() == "SUM":
            for cell in row:
                cell.font = bold_font

    workbook.save(buffer)
    return base64.b64encode(buffer.getvalue()).decode("ascii")


def _build_chapter_excel(chapter: pd.DataFrame) -> str:
    if chapter.empty:
        return ""

    buffer = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Kapitteloppsummering"

    headers = list(chapter.columns)
    worksheet.append(headers)

    header_fill = PatternFill(fill_type="solid", start_color="E2E8F0", end_color="E2E8F0")
    header_font = Font(bold=True, color="0F172A")
    header_alignment = Alignment(horizontal="center", vertical="center")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    string_columns = {"kapittel", "kapittel_navn", "laveste_tilbyder"}
    percent_columns = {"spann_pct"}
    numeric_columns = [col for col in headers if col not in string_columns.union(percent_columns)]
    for record in chapter.to_dict(orient="records"):
        row_values: list[Any] = []
        for column in headers:
            value = record.get(column)
            if isinstance(value, np.generic):
                value = value.item()
            if column in percent_columns and value not in (None, "", "SUM"):
                try:
                    row_values.append(float(value) / 100.0)
                    continue
                except (TypeError, ValueError):
                    pass
            if column in numeric_columns and value not in (None, "", "SUM"):
                try:
                    row_values.append(float(value))
                    continue
                except (TypeError, ValueError):
                    pass
            row_values.append(value)
        worksheet.append(row_values)

    worksheet.freeze_panes = "A2"

    border_side = Side(style="thin", color="CBD5F5")
    table_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = table_border

    for index, column in enumerate(headers, start=1):
        column_letter = get_column_letter(index)
        max_length = 0
        for cell in worksheet[column_letter]:
            if cell.value is None:
                continue
            length = len(str(cell.value))
            max_length = max(max_length, length)
        worksheet.column_dimensions[column_letter].width = max(12, max_length + 2)

        if column in numeric_columns:
            for cell in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=index, max_col=index):
                target = cell[0]
                if isinstance(target.value, (int, float)):
                    target.number_format = '#,##0.00'
                    target.alignment = Alignment(horizontal="right")
        elif column in percent_columns:
            for cell in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=index, max_col=index):
                target = cell[0]
                if isinstance(target.value, (int, float)):
                    target.number_format = '0.00%'
                    target.alignment = Alignment(horizontal="right")

    bold_font = Font(bold=True)
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        if str(row[0].value).strip().upper() == "SUM":
            for cell in row:
                cell.font = bold_font

    workbook.save(buffer)
    return base64.b64encode(buffer.getvalue()).decode("ascii")


def _format_parenthesized_currency(value: Any) -> str:
    try:
        num = float(value)
    except (TypeError, ValueError):
        return ""
    formatted = f"{num:,.2f}".replace(",", " ").replace(".", ",")
    return f"(kr {formatted})"


@app.post("/api/bid-compare")
async def bid_compare(files: list[UploadFile] = File(...)) -> dict[str, Any]:
    if not files:
        raise HTTPException(status_code=400, detail="No files uploaded.")

    bids: dict[str, pd.DataFrame] = {}
    errors: list[str] = []

    for upload in files:
        name = upload.filename or f"bid_{len(bids) + 1}"
        data = await upload.read()
        if not data:
            errors.append(f"{name} is empty.")
            continue
        try:
            if name.lower().endswith(".xml"):
                df_clean = _parse_ns3459_xml(data, name)
                try:
                    candidate = _extract_company_name(data) or name
                except Exception:
                    candidate = name
            else:
                df_raw = _read_tabular(name, data)
                df_clean = _normalize_columns(df_raw)
                candidate = name

            unique_name = candidate
            counter = 1
            while unique_name in bids:
                counter += 1
                unique_name = f"{candidate} ({counter})"

            bids[unique_name] = df_clean
        except HTTPException as exc:
            errors.append(exc.detail)
        except Exception as exc:
            errors.append(f"Could not read {name}: {exc}")

    if not bids:
        raise HTTPException(status_code=400, detail="Could not read any files.")

    # Prepare normalized data
    normalized = {name: _to_records(df) for name, df in bids.items()}

    # Comparison matrix
    matrix = pd.DataFrame(columns=["postnr"])
    provider_order: list[str] = []
    unit_columns: list[str] = []
    sum_columns: list[str] = []
    sum_column_provider: dict[str, str] = {}
    post_meta: dict[str, dict[str, Any]] = {}
    option_totals: dict[str, float] = {}
    base_bids: dict[str, pd.DataFrame] = {}

    for idx, (name, df) in enumerate(bids.items()):
        provider_order.append(name)
        df_base = df[df["is_option"] != True].copy()
        base_bids[name] = df_base
        option_totals[name] = float(df[df["is_option"] == True]["sum_amount"].sum())

        unit_col = f"{name} (enhetspris)"
        sum_col = f"{name} (sum)"
        part = _aggregate_bid_rows(df_base, unit_col, sum_col)
        if idx == 0:
            matrix = part
        else:
            matrix = matrix.merge(part, on="postnr", how="outer")
        if unit_col not in matrix.columns:
            matrix[unit_col] = 0.0
        if sum_col not in matrix.columns:
            matrix[sum_col] = 0.0
        unit_columns.append(unit_col)
        sum_columns.append(sum_col)
        sum_column_provider[sum_col] = name

        for _, row in df_base.iterrows():
            postnr = str(row.get("postnr") or "").strip()
            if not postnr:
                continue
            entry = post_meta.setdefault(
                postnr,
                {
                    "kapittel": "",
                    "kapittel_navn": "",
                    "ns_code": "",
                    "specification": "",
                    "enhet": "",
                    "qty": None,
                },
            )
            kapittel = str(row.get("kapittel") or "").strip()
            kapittel_navn = str(row.get("kapittel_navn") or "").strip()
            if kapittel and not entry["kapittel"]:
                entry["kapittel"] = kapittel
            if kapittel_navn and not entry["kapittel_navn"]:
                entry["kapittel_navn"] = kapittel_navn
            if not entry["ns_code"]:
                ns_code = str(row.get("ns_code") or "").strip()
                if ns_code:
                    entry["ns_code"] = ns_code
            if not entry["specification"]:
                specification = str(row.get("specification") or "").strip()
                if specification:
                    entry["specification"] = specification
            if not entry["enhet"]:
                enhet = str(row.get("enhet") or "").strip()
                if enhet:
                    entry["enhet"] = enhet
            qty_value = row.get("qty")
            if entry["qty"] in (None, 0, 0.0):
                qty_float = _to_float(qty_value)
                entry["qty"] = qty_float if entry["qty"] is None or qty_float else entry["qty"]

    if not matrix.empty:
        matrix["postnr"] = matrix["postnr"].astype(str)

    if post_meta:
        meta_df = pd.DataFrame.from_dict(post_meta, orient="index").reset_index().rename(
            columns={
                "index": "postnr",
                "kapittel": "kapittel",
                "kapittel_navn": "kapittel_navn",
                "ns_code": "ns_code",
                "specification": "specification",
                "enhet": "enhet",
                "qty": "qty",
            }
        )
        matrix = matrix.merge(meta_df, on="postnr", how="left")
        for column in ["ns_code", "specification", "enhet", "kapittel", "kapittel_navn"]:
            if column in matrix.columns:
                matrix[column] = matrix[column].fillna("").astype(str)
        if "qty" in matrix.columns:
            matrix["qty"] = pd.to_numeric(matrix["qty"], errors="coerce").fillna(0.0)

    base_columns = ["kapittel", "kapittel_navn", "postnr", "ns_code", "specification", "enhet", "qty"]

    if "ns_code" not in matrix.columns:
        matrix["ns_code"] = ""
    if "specification" not in matrix.columns:
        matrix["specification"] = ""
    if "enhet" not in matrix.columns:
        matrix["enhet"] = ""
    if "qty" not in matrix.columns:
        matrix["qty"] = 0.0

    provider_columns: list[str] = []
    for name in provider_order:
        unit_col = f"{name} (enhetspris)"
        sum_col = f"{name} (sum)"
        if unit_col in matrix.columns:
            provider_columns.append(unit_col)
        if sum_col in matrix.columns:
            provider_columns.append(sum_col)

    remaining_columns = [col for col in matrix.columns if col not in base_columns + provider_columns]
    ordered_columns = [col for col in base_columns if col in matrix.columns] + provider_columns + remaining_columns
    matrix = matrix[ordered_columns]

    active_sum_columns = [col for col in sum_columns if col in matrix.columns]
    active_unit_columns = [col for col in unit_columns if col in matrix.columns]

    if active_sum_columns:
        sums_df = matrix[active_sum_columns].apply(pd.to_numeric, errors="coerce")
        min_indices = sums_df.idxmin(axis=1, skipna=True)
        matrix["vinner"] = min_indices.map(sum_column_provider).fillna("")
        matrix["lavest_sum"] = sums_df.min(axis=1, skipna=True).fillna(0.0)
        std_series = sums_df.std(axis=1, skipna=True)
        matrix["std_avvik"] = std_series.fillna(0.0)
        snitt_series = sums_df.mean(axis=1, skipna=True)
        matrix["snitt"] = snitt_series.where(~snitt_series.eq(0.0), pd.NA)
        snitt_divisor = snitt_series.replace({0.0: np.nan})
        std_ratio = std_series.divide(snitt_divisor)
        std_ratio = std_ratio.replace([np.inf, -np.inf], np.nan).fillna(0.0)
        matrix["std_pct"] = (std_ratio * 100.0).astype(float)

        # Calculate z-scores for each bid
        num_bids = len(active_sum_columns)
        if num_bids >= 3:
            for sum_col in active_sum_columns:
                provider_name = sum_column_provider.get(sum_col, "")
                z_col = f"{provider_name} (z-score)"
                z_scores = (sums_df[sum_col] - snitt_series) / std_series
                z_scores = z_scores.replace([np.inf, -np.inf], np.nan).fillna(0.0)
                matrix[z_col] = z_scores.astype(float)
    else:
        matrix["vinner"] = ""
        matrix["lavest_sum"] = 0.0
        matrix["std_avvik"] = 0.0
        matrix["snitt"] = pd.NA
        matrix["std_pct"] = 0.0

    for column in ("kapittel", "kapittel_navn", "vinner"):
        if column in matrix.columns:
            matrix[column] = matrix[column].fillna("")

    if not matrix.empty:
        matrix = matrix.sort_values(by="postnr")

    if not matrix.empty:
        matrix_sum = {}
        z_score_columns = [col for col in matrix.columns if col.endswith("(z-score)")]

        for key in matrix.columns:
            if key == "postnr":
                matrix_sum[key] = "SUM"
            elif key in {"kapittel", "kapittel_navn", "vinner"}:
                matrix_sum[key] = ""
            elif key in active_sum_columns + ["lavest_sum", "std_avvik"]:
                matrix_sum[key] = float(matrix[key].sum(skipna=True))
            elif key in active_unit_columns:
                matrix_sum[key] = ""
            elif key in {"snitt", "std_pct"} or key in z_score_columns:
                matrix_sum[key] = ""
            else:
                matrix_sum[key] = ""
        matrix_disp = pd.concat([matrix, pd.DataFrame([matrix_sum])], ignore_index=True)
    else:
        matrix_disp = matrix.copy()

    palette = ["#BFDBFE", "#FDE68A", "#E9D5FF", "#BBF7D0", "#FBCFE8", "#FECACA", "#C7D2FE"]
    column_colors: dict[str, str] = {}
    for idx, name in enumerate(provider_order):
        base_color = palette[idx % len(palette)]
        unit_col = f"{name} (enhetspris)"
        sum_col = f"{name} (sum)"
        column_colors[unit_col] = base_color
        column_colors[sum_col] = base_color

    chapter_titles = _collect_chapter_titles(bids)

    # Chapter summary
    chapter = pd.DataFrame(columns=["kapittel"])
    for name, df in base_bids.items():
        if df.empty:
            continue
        part = df.groupby("kapittel", as_index=False)["sum_amount"].sum().rename(columns={"sum_amount": name})
        if chapter.empty:
            chapter = part
        else:
            chapter = chapter.merge(part, on="kapittel", how="outer")
    chapter = chapter.fillna(0.0)
    if not chapter.empty:
        chapter["kapittel"] = chapter["kapittel"].astype(str).str.strip()
        chapter["kapittel_navn"] = chapter["kapittel"].map(lambda code: chapter_titles.get(code, ""))
        provider_columns = [col for col in chapter.columns if col not in {"kapittel", "kapittel_navn"}]
        chapter["laveste_tilbyder"] = ""
        chapter["laveste_sum"] = 0.0
        chapter["spann_pct"] = 0.0
        for idx, row in chapter.iterrows():
            sums = {col: float(row[col]) for col in provider_columns if col in row}
            if not sums:
                continue
            min_value = min(sums.values())
            max_value = max(sums.values())
            if all(abs(v - min_value) < 1e-6 for v in sums.values()):
                chapter.at[idx, "laveste_tilbyder"] = "N/A"
                chapter.at[idx, "laveste_sum"] = 0.0
                chapter.at[idx, "spann_pct"] = 0.0
            else:
                min_provider = min(sums, key=sums.get)
                chapter.at[idx, "laveste_tilbyder"] = min_provider
                chapter.at[idx, "laveste_sum"] = float(min_value)
                chapter.at[idx, "spann_pct"] = float(0.0 if min_value == 0 else ((max_value - min_value) / min_value) * 100.0)

        ordered_cols = [
            "kapittel",
            "kapittel_navn",
            "laveste_tilbyder",
            "laveste_sum",
            "spann_pct",
        ] + [col for col in provider_columns if col in chapter.columns]
        chapter = chapter[ordered_cols]
        chapter_sum = {
            key: (
                "SUM"
                if key == "kapittel"
                else "" if key == "kapittel_navn"
                else "" if key == "laveste_tilbyder"
                else float(chapter[key].sum()) if key == "laveste_sum"
                else "" if key == "spann_pct"
                else chapter[key].sum()
            )
            for key in chapter.columns
        }
        chapter_disp = pd.concat([chapter, pd.DataFrame([chapter_sum])], ignore_index=True)
    else:
        chapter_disp = chapter.copy()

    # Summary metrics
    totals = {name: float(df["sum_amount"].sum()) for name, df in base_bids.items()} if base_bids else {}
    option_totals = {name: float(value) for name, value in option_totals.items()}
    if totals:
        min_total = min(totals.values())
        max_total = max(totals.values())
        if abs(max_total - min_total) < 1e-6:
            winner_name = ""
            winner_total = 0.0
        else:
            winner_name = min(totals, key=totals.get)
            winner_total = totals.get(winner_name, 0.0)
    else:
        winner_name = ""
        winner_total = 0.0
    if base_bids:
        post_count = int(
            len(set(pd.concat([df["postnr"] for df in base_bids.values()], ignore_index=True)))
        )
    else:
        post_count = 0

    # Build Excel output
    excel_buffer = io.BytesIO()
    with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
        for name, df in bids.items():
            sheet_df = df.copy()
            option_mask = sheet_df.get("is_option") == True
            if isinstance(option_mask, pd.Series) and option_mask.any():
                for col in ["unit_price", "sum_amount"]:
                    if col in sheet_df.columns:
                        sheet_df[col] = sheet_df[col].astype(object)
                        formatted_values = sheet_df.loc[option_mask, col].apply(
                            lambda x: _format_parenthesized_currency(x) if pd.notna(x) else ""
                        )
                        sheet_df.loc[option_mask, col] = formatted_values
            if "is_option" in sheet_df.columns:
                sheet_df = sheet_df.drop(columns=["is_option"])
            sheet_df.to_excel(writer, index=False, sheet_name=name[:28])
        base_matrix = matrix.copy()
        if "is_option" in base_matrix.columns:
            base_matrix = base_matrix.drop(columns=["is_option"])
        base_matrix.to_excel(writer, index=False, sheet_name="Sammenligning")
        chapter.to_excel(writer, index=False, sheet_name="Kapittel")
    excel_bytes = excel_buffer.getvalue()
    excel_b64 = base64.b64encode(excel_bytes).decode("ascii")

    matrix_excel = _build_matrix_excel(matrix_disp, active_sum_columns, active_unit_columns, column_colors)
    chapter_excel = _build_chapter_excel(chapter_disp)

    result = {
        "normalized": normalized,
        "matrix": {
            "columns": list(matrix_disp.columns),
            "rows": _to_records(matrix_disp),
        },
        "chapters": {
            "columns": list(chapter_disp.columns),
            "rows": _to_records(chapter_disp),
        },
        "summary": {
            "totals": totals,
            "option_totals": option_totals,
            "winner": {"name": winner_name, "total": winner_total},
            "post_count": post_count,
        },
        "generated_at": datetime.now().isoformat(),
        "excel": excel_b64,
        "matrix_excel": matrix_excel,
        "chapters_excel": chapter_excel,
        "errors": errors,
    }
    return result


@app.get("/health")
async def health():
    return {"status": "healthy"}
